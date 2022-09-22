import openpyxl


file="C:\SeleniumPractice\data.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row	#count no of rows in a excel sheet
cols=sheet.max_column	#count no of columns in a excel sheet 

# Reading all the rows and collumns from excel sheet

for r in range(1,rows+1):
	for c in range(1,cols+1):
		print(sheet.cell(r,c).value, end='     ')
	print()