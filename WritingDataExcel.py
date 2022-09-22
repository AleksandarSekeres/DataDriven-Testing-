# writing same value data in all collumns
import openpyxl

file="C:\\SeleniumPractice\\testdata.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active # or sheet=workbook["Data"]  -- get active sheet from excel

for r in range(1,6):
	for c in range(1,4):
		sheet.cell(r,c).value="welcome"

workbook.save(file)


# multiple data, if the data is different we cannot put it in loop 


file="C:\\SeleniumPractice\\testdata.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook.active # or sheet=workbook["Data"]  

sheet.cell(1,1).value="welcome"
sheet.cell(1,2).value="123

sheet.cell(2,1).value="world"
sheet.cell(2,2).value=789

workbook.save(file)